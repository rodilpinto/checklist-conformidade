# -*- coding: utf-8 -*-
"""
Gerador de planilha Excel formatada para checklists de conformidade normativa.

Produz um arquivo .xlsx em memória (bytes) pronto para download via Streamlit,
com estilos padronizados, validação de dados, auto-filtro e freeze panes.

Dependências: openpyxl (>= 3.1)

Uso:
    from lib.excel_builder import build_excel

    items = [
        {
            "id": 1,
            "capitulo": "Cap. I",
            "artigo": "Art. 1º",
            "texto_literal": "Texto do dispositivo...",
            "requisito": "Descrição do requisito...",
            "risco": "Descrição do risco...",
            "nivel": "Alto",
            "mitigacao": "Ação de mitigação...",
            "responsavel": "Gestor de Negócio",
            "evidencia": "Documento comprobatório...",
        },
        ...
    ]
    xlsx_bytes = build_excel(items, title="Checklist de Conformidade")
    st.download_button("Baixar Excel", xlsx_bytes, "checklist.xlsx")
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTES DE ESTILO
# Padrão visual reutilizado dos scripts create_v106.py e
# create_checklist_roteiro_levantamento.py do projeto.
# ═══════════════════════════════════════════════════════════════════════════════

# -- Fontes -----------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
DATA_FONT_BOLD = Font(name="Arial", size=10, bold=True)
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F4E79")
RISK_FONT_WHITE = Font(name="Arial", size=10, bold=True, color="FFFFFF")
RISK_FONT_DARK = Font(name="Arial", size=10, bold=True, color="000000")

# -- Preenchimentos ----------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F7FB")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

# -- Cores por nível de risco (MCGR - Câmara dos Deputados) -----------------
RISK_FILLS: dict[str, PatternFill] = {
    "Muito Alto": PatternFill("solid", fgColor="FF4444"),   # vermelho
    "Alto": PatternFill("solid", fgColor="FFA500"),          # laranja
    "Moderado": PatternFill("solid", fgColor="FFD700"),      # amarelo
    "Baixo": PatternFill("solid", fgColor="92D050"),         # verde
}

# "Muito Alto" e "Alto" usam fonte branca para contraste sobre fundo
# escuro (vermelho/laranja); "Moderado" e "Baixo" usam fonte escura.
RISK_FONTS: dict[str, Font] = {
    "Muito Alto": RISK_FONT_WHITE,
    "Alto": RISK_FONT_WHITE,
    "Moderado": RISK_FONT_DARK,
    "Baixo": RISK_FONT_DARK,
}

# -- Bordas ------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# -- Alinhamentos ------------------------------------------------------------
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# -- Layout de colunas -------------------------------------------------------
# (chave do dict, título do header, largura em caracteres)
COLUMNS: list[tuple[str, str, int]] = [
    ("id",              "Nº",              5),
    ("capitulo",        "Capítulo",       14),
    ("artigo",          "Artigo",         12),
    ("texto_literal",   "Texto Literal",  55),
    ("requisito",       "Requisito",      40),
    ("risco",           "Risco",          40),
    ("probabilidade",   "Prob.",           7),
    ("impacto",         "Impacto",         8),
    ("nivel",           "Nível",          12),
    ("mitigacao",       "Mitigação",      40),
    ("responsavel",     "Responsável",    22),
    ("evidencia",       "Evidência",      30),
    ("status",          "Status",         16),
    ("observacoes",     "Observações",    30),
]

# Valores permitidos para validação de dados
STATUS_OPTIONS = "Não Iniciado,Em Andamento,Concluído,Não Aplicável"
NIVEL_OPTIONS = "Muito Alto,Alto,Moderado,Baixo"

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS INTERNOS
# ═══════════════════════════════════════════════════════════════════════════════


def _style_header_row(ws: Worksheet, row: int, num_cols: int) -> None:
    """Aplica estilo de cabeçalho (azul escuro, fonte branca) a uma linha."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER
        cell.border = THIN_BORDER


def _style_section_row(ws: Worksheet, row: int, num_cols: int, label: str) -> None:
    """Insere linha de separação visual para mudança de capítulo (sem merge)."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER
        ws.cell(row=row, column=col_idx).fill = SECTION_FILL


def _apply_risk_style(cell: Any, nivel: str) -> None:
    """Aplica cor de fundo e fonte ao campo Nível conforme classificação de risco."""
    nivel_normalizado = (nivel or "").strip()
    if nivel_normalizado in RISK_FILLS:
        cell.fill = RISK_FILLS[nivel_normalizado]
        cell.font = RISK_FONTS[nivel_normalizado]


def _auto_fit_row_heights(
    ws: Worksheet,
    header_row: int,
    next_empty_row: int,
    columns: list[tuple[str, str, int]],
) -> None:
    """Ajusta a altura de cada linha proporcionalmente ao texto mais longo.

    Estima quantas linhas visuais o texto ocupa dentro da largura da coluna
    e define a altura da linha de forma que o conteúdo fique visível.
    """
    _CHAR_WIDTH_FACTOR = 0.85  # caracteres por unidade de largura Excel (Arial 10pt)
    _LINE_HEIGHT = 15  # pontos por linha de texto
    _MIN_HEIGHT = 30
    _MAX_HEIGHT = 300
    _HEADER_HEIGHT = 36

    ws.row_dimensions[header_row].height = _HEADER_HEIGHT

    col_widths = [w for (_, _, w) in columns]

    for row_idx in range(header_row + 1, next_empty_row):
        max_lines = 1
        for col_idx, width in enumerate(col_widths, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            text = str(cell.value or "")
            if not text:
                continue
            chars_per_line = max(int(width * _CHAR_WIDTH_FACTOR), 1)
            # Conta quebras de linha explícitas + wrapping estimado
            lines = 0
            for paragraph in text.split("\n"):
                lines += max(1, -(-len(paragraph) // chars_per_line))  # ceil division
            max_lines = max(max_lines, lines)

        height = max(_MIN_HEIGHT, min(_MAX_HEIGHT, max_lines * _LINE_HEIGHT))
        ws.row_dimensions[row_idx].height = height


def _build_legend_sheet(wb: Workbook) -> None:
    """Cria a aba 'Legenda' com critérios de risco e aviso sobre IA."""
    ws = wb.create_sheet("Legenda")

    # Larguras
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    row = 1

    # Título
    cell = ws.cell(row=row, column=1, value="Legenda e Critérios de Avaliação")
    cell.font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36
    row += 2

    # Seção: Metodologia de Gestão de Riscos
    cell = ws.cell(row=row, column=1, value="Metodologia de Gestão de Riscos")
    cell.font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    row += 1

    cell = ws.cell(row=row, column=1, value=(
        "Baseada no Modelo Corporativo de Gestão de Riscos (MCGR) da Câmara dos "
        "Deputados (Ato da Mesa nº 233/2018), que adota referências da ABNT NBR "
        "ISO 31000, COSO-ERM e PMI."
    ))
    cell.font = Font(name="Arial", size=10, italic=True)
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30
    row += 2

    # Subseção: Escala de Probabilidade
    cell = ws.cell(row=row, column=1, value="Escala de Probabilidade")
    cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    row += 1

    prob_scale = [
        ("5 — Praticamente certo", "Ocorrência quase garantida no prazo associado ao escopo."),
        ("4 — Muito provável", "Repete-se com elevada frequência ou há muitos indícios de que ocorrerá."),
        ("3 — Provável", "Repete-se com frequência razoável ou há indícios de que possa ocorrer."),
        ("2 — Pouco provável", "Histórico aponta para baixa frequência de ocorrência."),
        ("1 — Raro", "Acontece apenas em situações excepcionais; sem histórico conhecido."),
    ]

    for label, desc in prob_scale:
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = Font(name="Arial", size=10, bold=True)
        cell_a.border = THIN_BORDER
        cell_a.alignment = Alignment(vertical="center")
        cell_b = ws.cell(row=row, column=2, value=desc)
        cell_b.font = Font(name="Arial", size=10)
        cell_b.alignment = ALIGN_WRAP
        cell_b.border = THIN_BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    # Subseção: Escala de Impacto
    cell = ws.cell(row=row, column=1, value="Escala de Impacto")
    cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    row += 1

    impact_scale = [
        ("5 — Muito alto", "Compromete totalmente ou quase totalmente o atingimento do objetivo."),
        ("4 — Alto", "Compromete a maior parte do atingimento do objetivo."),
        ("3 — Médio", "Compromete razoavelmente o atingimento do objetivo."),
        ("2 — Baixo", "Compromete em alguma medida o alcance do objetivo."),
        ("1 — Muito baixo", "Compromete minimamente ou não altera o atingimento do objetivo."),
    ]

    for label, desc in impact_scale:
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = Font(name="Arial", size=10, bold=True)
        cell_a.border = THIN_BORDER
        cell_a.alignment = Alignment(vertical="center")
        cell_b = ws.cell(row=row, column=2, value=desc)
        cell_b.font = Font(name="Arial", size=10)
        cell_b.alignment = ALIGN_WRAP
        cell_b.border = THIN_BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    # Subseção: Níveis de Risco (Criticidade = Probabilidade × Impacto)
    cell = ws.cell(row=row, column=1, value="Níveis de Risco (Criticidade = P × I)")
    cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    row += 1

    risk_levels = [
        ("Muito Alto", "FF4444", "FFFFFF",
         "Criticidade 20 a 25 — Risco inaceitável que exige tratamento "
         "imediato. Pode comprometer totalmente o atingimento dos objetivos."),
        ("Alto", "FFA500", "FFFFFF",
         "Criticidade 10 a 16 — Risco significativo que demanda ações "
         "prioritárias de tratamento para reduzir a exposição."),
        ("Moderado", "FFD700", "000000",
         "Criticidade 4 a 9 — Risco tolerável sob monitoramento. Pode "
         "exigir ações de mitigação conforme o apetite a riscos definido."),
        ("Baixo", "92D050", "000000",
         "Criticidade 1 a 3 — Risco aceitável. Geralmente aceito sem "
         "necessidade de tratamento adicional."),
    ]

    for nivel, bg_color, fg_color, descricao in risk_levels:
        cell_nivel = ws.cell(row=row, column=1, value=nivel)
        cell_nivel.font = Font(name="Arial", size=11, bold=True, color=fg_color)
        cell_nivel.fill = PatternFill("solid", fgColor=bg_color)
        cell_nivel.alignment = Alignment(horizontal="center", vertical="center")
        cell_nivel.border = THIN_BORDER

        cell_desc = ws.cell(row=row, column=2, value=descricao)
        cell_desc.font = Font(name="Arial", size=10)
        cell_desc.alignment = ALIGN_WRAP
        cell_desc.border = THIN_BORDER
        ws.row_dimensions[row].height = 45
        row += 1

    row += 1

    # Seção: Aviso sobre IA
    cell = ws.cell(row=row, column=1, value="Aviso Importante")
    cell.font = Font(name="Arial", size=12, bold=True, color="CC0000")
    row += 1

    aviso = (
        "Esta planilha foi gerada automaticamente por inteligência artificial "
        "(Google Gemini) a partir do texto do normativo informado.\n\n"
        "A classificação de risco (Muito Alto, Alto, Moderado, Baixo), bem "
        "como os valores de probabilidade e impacto, são uma SUGESTÃO INICIAL "
        "produzida pela IA com base no teor do dispositivo legal e na "
        "metodologia MCGR da Câmara dos Deputados. Ela NÃO substitui o "
        "julgamento profissional do auditor, gestor ou responsável pela "
        "conformidade.\n\n"
        "É responsabilidade do usuário que gera e utiliza esta planilha:\n"
        "  • Revisar todos os itens e suas classificações;\n"
        "  • Ajustar os níveis de risco conforme o contexto organizacional;\n"
        "  • Validar os requisitos contra o texto original do normativo;\n"
        "  • Complementar ou remover itens conforme necessário.\n\n"
        "A ferramenta é um auxílio para acelerar o trabalho — a decisão "
        "final e a responsabilidade são sempre do profissional."
    )
    cell_aviso = ws.cell(row=row, column=1, value=aviso)
    cell_aviso.font = Font(name="Arial", size=10)
    cell_aviso.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell_aviso.border = THIN_BORDER
    ws.row_dimensions[row].height = 200

    row += 2

    # Seção: Campos da planilha
    cell = ws.cell(row=row, column=1, value="Campos da Planilha")
    cell.font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    row += 1

    campos = [
        ("Nº", "Número sequencial do item."),
        ("Capítulo", "Capítulo ou seção do normativo."),
        ("Artigo", "Artigo, inciso, parágrafo ou alínea específica."),
        ("Texto Literal", "Transcrição literal do dispositivo legal (sem paráfrase)."),
        ("Requisito", "O que deve ser verificado ou atendido."),
        ("Risco", "Consequência do não atendimento ao requisito."),
        ("Prob.", "Probabilidade de ocorrência (1-5, conforme escala MCGR)."),
        ("Impacto", "Impacto sobre os objetivos (1-5, conforme escala MCGR)."),
        ("Nível", "Criticidade (P×I): Muito Alto (20-25), Alto (10-16), Moderado (4-9), Baixo (1-3)."),
        ("Mitigação", "Ação sugerida para atender ao requisito (evitar, transferir, mitigar ou aceitar)."),
        ("Responsável", "Ator ou área responsável pelo atendimento."),
        ("Evidência", "Documento ou artefato que comprova o atendimento."),
        ("Status", "Andamento: Não Iniciado, Em Andamento, Concluído, Não Aplicável."),
        ("Observações", "Notas livres do avaliador."),
    ]

    for campo, descricao in campos:
        cell_campo = ws.cell(row=row, column=1, value=campo)
        cell_campo.font = Font(name="Arial", size=10, bold=True)
        cell_campo.border = THIN_BORDER
        cell_campo.alignment = Alignment(vertical="center")

        cell_desc = ws.cell(row=row, column=2, value=descricao)
        cell_desc.font = Font(name="Arial", size=10)
        cell_desc.border = THIN_BORDER
        cell_desc.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1


def _safe_value(value: Any) -> Any:
    """Retorna valor seguro para célula Excel, tratando None.

    Preserva int/float para que o Excel os reconheça como números;
    converte o restante para string.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value)


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════


def build_excel(
    items: list[dict],
    title: str = "Checklist de Conformidade",
) -> bytes:
    """
    Gera planilha Excel formatada a partir de uma lista de itens de checklist.

    Parameters
    ----------
    items : list[dict]
        Lista de dicionários, cada um com as chaves:
        id, capitulo, artigo, texto_literal, requisito, risco, nivel,
        mitigacao, responsavel, evidencia.
        As chaves ``status`` e ``observacoes`` são opcionais (default: vazio).

    title : str
        Título exibido na aba da planilha (máx. 31 caracteres, limitação Excel).

    Returns
    -------
    bytes
        Conteúdo do arquivo .xlsx pronto para download (BytesIO.getvalue()).

    Raises
    ------
    ValueError
        Se ``items`` estiver vazio.

    Example
    -------
    >>> data = [{"id": 1, "capitulo": "Cap. I", "artigo": "Art. 1º",
    ...          "texto_literal": "...", "requisito": "...", "risco": "...",
    ...          "nivel": "Alto", "mitigacao": "...", "responsavel": "...",
    ...          "evidencia": "..."}]
    >>> xlsx = build_excel(data, title="Meu Checklist")
    >>> isinstance(xlsx, bytes)
    True
    """
    if not items:
        raise ValueError("A lista de itens não pode estar vazia.")

    wb = Workbook()
    ws = wb.active

    # Título da aba (Excel limita a 31 caracteres e proíbe \/:*?[])
    safe_title = title
    for ch in ("\\", "/", ":", "*", "?", "[", "]"):
        safe_title = safe_title.replace(ch, "_")
    ws.title = safe_title[:31]

    num_cols = len(COLUMNS)

    # ── 1. Definir larguras das colunas ──────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 2. Escrever cabeçalhos ───────────────────────────────────────────────
    header_row = 1
    for col_idx, (_, header_text, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=header_row, column=col_idx, value=header_text)

    _style_header_row(ws, header_row, num_cols)

    # ── 3. Auto-filtro na linha de cabeçalho ─────────────────────────────────
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"

    # ── 4. Freeze panes (fixar cabeçalho ao rolar) ──────────────────────────
    ws.freeze_panes = "A2"

    # ── 5. Índice da coluna "Nível" (para estilo de risco) ──────────────────
    nivel_col_idx = next(
        i for i, (key, _, _) in enumerate(COLUMNS, start=1) if key == "nivel"
    )
    status_col_idx = next(
        i for i, (key, _, _) in enumerate(COLUMNS, start=1) if key == "status"
    )

    # ── 6. Escrever dados com separadores de capítulo ────────────────────────
    current_row = header_row + 1
    previous_capitulo: str | None = None
    data_row_count = 0  # conta apenas linhas de dados (não separadores)

    for item in items:
        capitulo = str(item.get("capitulo", "") or "")

        # Inserir separador visual quando o capítulo muda
        if capitulo and capitulo != previous_capitulo:
            _style_section_row(ws, current_row, num_cols, capitulo)
            current_row += 1
            previous_capitulo = capitulo

        # Determinar preenchimento de fundo alternado
        is_odd = data_row_count % 2 == 1
        row_fill = ALT_ROW_FILL if is_odd else WHITE_FILL

        # Escrever cada célula da linha
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = _safe_value(item.get(key, ""))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill

            # Colunas curtas (Nº, Nível, Status) ficam centralizadas;
            # demais com wrap_text alinhado ao topo.
            if key in ("id", "probabilidade", "impacto", "nivel", "status"):
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_WRAP

        # Aplicar estilo de cor ao campo Nível
        nivel_cell = ws.cell(row=current_row, column=nivel_col_idx)
        _apply_risk_style(nivel_cell, str(item.get("nivel", "") or ""))

        current_row += 1
        data_row_count += 1

    last_data_row = current_row - 1

    # ── 7. Atualizar referência do auto-filtro para incluir todas as linhas ──
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"

    # ── 8. Data validation: Status ───────────────────────────────────────────
    #    Aplica a todas as células de dados na coluna Status.
    status_col_letter = get_column_letter(status_col_idx)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Selecione: Não Iniciado, Em Andamento, Concluído ou Não Aplicável.",
        showInputMessage=True,
        promptTitle="Status",
        prompt="Selecione o status do item.",
    )
    dv_status.add(f"{status_col_letter}2:{status_col_letter}{last_data_row}")
    ws.add_data_validation(dv_status)

    # ── 9. Data validation: Nível ────────────────────────────────────────────
    nivel_col_letter = get_column_letter(nivel_col_idx)
    dv_nivel = DataValidation(
        type="list",
        formula1=f'"{NIVEL_OPTIONS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Selecione: Muito Alto, Alto, Moderado ou Baixo.",
        showInputMessage=True,
        promptTitle="Nível de Risco",
        prompt="Selecione a classificação de risco.",
    )
    dv_nivel.add(f"{nivel_col_letter}2:{nivel_col_letter}{last_data_row}")
    ws.add_data_validation(dv_nivel)

    # ── 10. Auto-ajuste de altura das linhas conforme conteúdo ───────────────
    _auto_fit_row_heights(ws, header_row, current_row, COLUMNS)

    # ── 11. Configuração de impressão (paisagem, ajustar à largura) ─────────
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # ── 12. Criar aba de legenda ──────────────────────────────────────────────
    _build_legend_sheet(wb)

    # ── 13. Salvar em memória e retornar bytes ────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
